"""Create an Excel report for recent patent-writing feedback fixes.

This script uses only Python standard library modules so it works in the
current lightweight project environment without extra spreadsheet packages.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import html
import sys


OUTPUT = Path("outputs/patent_feedback_fix_report.xlsx")


FIX_ROWS = [
    [
        "1",
        "标题仍然过长，出现“基于……的……”句式",
        "把短标题从提示建议升级为硬规则，并新增标题清洗兜底逻辑。",
        "代码 + 提示词 + Skill",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md",
        "候选生成提示禁止“基于……的……”；_clean_candidate_title() 会把“基于X的Y”清洗成“Y”。",
        "示例已验证：基于VMD频带能量特征与时序网络的铝电解槽阳极效应早期预警方法 -> 铝电解槽阳极效应早期预警方法。",
    ],
    [
        "2",
        "简短标题，去掉细节，只在标题表现发明是什么",
        "更新最终文档结构和每章写作提示，要求标题页和发明名称只保留发明对象。",
        "提示词 + 章节结构",
        "patent_discovery_agent.py",
        "FINAL_FORMAT_GUIDE、WRITING_STEPS、_generate_section()、_revise_section() 均加入短标题约束。",
        "后续标题页、发明名称、候选专利名称都会受到同一约束。",
    ],
    [
        "3",
        "增加发明的目的",
        "把“发明目的”加入发明内容的固定章节顺序。",
        "章节结构 + Skill",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md；writer.py",
        "四、发明内容中新增“发明目的”，位于关键创新点之后、拟解决技术问题之前。",
        "交互式写作和旧模板写作都会生成该部分。",
    ],
    [
        "4",
        "增加保护范围",
        "新增独立章节“保护范围”。",
        "章节结构 + 路由",
        "patent_discovery_agent.py；skill_router.py；writer.py；skills/patent-writing/SKILL.md",
        "最终结构新增“五、保护范围”，要求覆盖方法、系统、装置、存储介质及核心技术特征边界。",
        "章节序号同步调整为八个主章节。",
    ],
    [
        "5",
        "出现两处技术交底书，去掉后面那个",
        "取消组装最终 Markdown 时额外追加候选标题，避免标题页之外再重复生成文档标题。",
        "代码",
        "patent_discovery_agent.py",
        "_assemble_document() 不再自动插入 '# candidate.title'，只拼接已确认章节。",
        "标题来源集中在标题页和发明名称章节，减少重复。",
    ],
    [
        "6",
        "背景技术过于冗余拖沓",
        "把背景技术约束为2-3个与本发明直接相关的问题点。",
        "提示词 + Skill + 模板",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md；writer.py",
        "禁止泛泛行业综述；禁止使用“技术空白一/二/三”“补足技术空白”等直白表达。",
        "背景技术会更接近“行业内有a/b/c问题”的写法。",
    ],
    [
        "7",
        "不要把“补技术空白”直白写成创新点",
        "把相关表达列为禁止项，并要求用具体技术差异表达创新。",
        "提示词 + Skill",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md",
        "明确禁止断言“整个行业完全没有某项技术”，禁止直接写“补了技术空白”。",
        "降低专利文本中过度绝对化、审查风险较高的表述。",
    ],
    [
        "8",
        "拟解决的技术问题要和背景问题对应",
        "在生成与修改提示中加入一一对应规则。",
        "提示词",
        "patent_discovery_agent.py",
        "要求背景提出a/b/c，发明内容就针对a/b/c解决。",
        "减少背景和发明内容脱节。",
    ],
    [
        "9",
        "关键创新点不要作为第十一部分单开再重复",
        "把关键创新点移到发明内容开头，并从最终结构中删除单独的第十一部分。",
        "章节结构 + Skill",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md",
        "四、发明内容现在从“关键创新点”开始，后续不再生成“区别于现有技术的关键创新点”单独章节。",
        "避免同一内容重复出现。",
    ],
    [
        "10",
        "有益效果的数据来源不明确，不能乱写量化结果",
        "加入证据约束：有材料支撑才写量化，无支撑只写简洁定性效果。",
        "提示词 + Skill + 模板",
        "patent_discovery_agent.py；skills/patent-writing/SKILL.md；writer.py",
        "禁止编造百分比、金额、精度提升等；旧模板也改为“现有材料尚未提供明确实验数据时仅作定性说明”。",
        "避免生成看似精确但没有依据的数据。",
    ],
    [
        "11",
        "修改意见后重写也必须遵守这些要求",
        "同步更新章节修改函数的提示词。",
        "提示词",
        "patent_discovery_agent.py",
        "_revise_section() 与 _generate_section() 使用同一套标题、背景、创新点、有益效果约束。",
        "用户选择“提修改意见”后不会绕过规则。",
    ],
]


FILE_ROWS = [
    [
        "patent_discovery_agent.py",
        "核心交互式 agent 写作逻辑",
        "更新最终文档结构、写作步骤、候选生成提示、章节生成提示、章节修改提示、标题清洗函数、最终文档组装逻辑。",
    ],
    [
        "skills/patent-writing/SKILL.md",
        "项目内专利写作 skill",
        "同步新的标准结构和写作约束，明确短标题、背景聚焦、问题对应、有益效果证据化。",
    ],
    [
        "writer.py",
        "旧模板写作兜底",
        "把旧版固定模板改为新结构，新增发明目的和保护范围，去掉硬编码场景和重复标题。",
    ],
    [
        "skill_router.py",
        "任务路由与章节清单",
        "同步技术方案文档章节清单，新增保护范围，调整输出标题和建议检索词。",
    ],
]


def col_name(index: int) -> str:
    name = ""
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def cell_xml(row: int, col: int, value: str, style: int = 0) -> str:
    ref = f"{col_name(col)}{row}"
    escaped = html.escape(str(value), quote=False)
    return f'<c r="{ref}" t="inlineStr" s="{style}"><is><t>{escaped}</t></is></c>'


def sheet_xml(title: str, headers: list[str], rows: list[list[str]]) -> str:
    all_rows = [[title], [], headers, *rows]
    xml_rows = []
    for r_idx, row in enumerate(all_rows, start=1):
        cells = []
        for c_idx, value in enumerate(row, start=1):
            style = 1 if r_idx == 1 else 2 if r_idx == 3 else 0
            cells.append(cell_xml(r_idx, c_idx, value, style))
        xml_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')

    cols = "".join(
        f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
        for idx, width in enumerate([8, 32, 38, 20, 38, 50, 42][: len(headers)], start=1)
    )
    max_row = len(all_rows)
    max_col = col_name(len(headers))
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="A1:{max_col}{max_row}"/>
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="3" topLeftCell="A4" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <sheetFormatPr defaultRowHeight="18"/>
  <cols>{cols}</cols>
  <sheetData>{"".join(xml_rows)}</sheetData>
  <autoFilter ref="A3:{max_col}{max_row}"/>
  <pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>
</worksheet>'''


def build_xlsx() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    if _build_with_openpyxl():
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sheet1 = sheet_xml(
        f"专利写作反馈修复对照表（生成时间：{now}）",
        ["序号", "用户反馈 / 问题", "修复策略", "修复类型", "涉及文件", "关键实现", "效果 / 验证"],
        FIX_ROWS,
    )
    sheet2 = sheet_xml(
        "文件变更摘要",
        ["文件", "作用", "本次改动摘要"],
        FILE_ROWS,
    )

    with ZipFile(OUTPUT, "w", ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", CONTENT_TYPES)
        zf.writestr("_rels/.rels", RELS)
        zf.writestr("docProps/core.xml", CORE_PROPS)
        zf.writestr("docProps/app.xml", APP_PROPS)
        zf.writestr("xl/workbook.xml", WORKBOOK)
        zf.writestr("xl/_rels/workbook.xml.rels", WORKBOOK_RELS)
        zf.writestr("xl/styles.xml", STYLES)
        zf.writestr("xl/worksheets/sheet1.xml", sheet1)
        zf.writestr("xl/worksheets/sheet2.xml", sheet2)


def _build_with_openpyxl() -> bool:
    local_target = Path("/private/tmp/openpyxl_agent_report")
    if local_target.exists():
        sys.path.insert(0, str(local_target))
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "修复对照"
    summary = wb.create_sheet("文件摘要")

    title_fill = PatternFill("solid", fgColor="145A4A")
    header_fill = PatternFill("solid", fgColor="28715F")
    white_bold = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=16)
    border = Border(
        left=Side(style="thin", color="D9E2DD"),
        right=Side(style="thin", color="D9E2DD"),
        top=Side(style="thin", color="D9E2DD"),
        bottom=Side(style="thin", color="D9E2DD"),
    )
    body_alignment = Alignment(vertical="top", wrap_text=True)

    def write_sheet(sheet, title: str, headers: list[str], rows: list[list[str]], widths: list[int]) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        sheet.cell(1, 1, title)
        sheet.cell(1, 1).fill = title_fill
        sheet.cell(1, 1).font = title_font
        sheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28
        sheet.append([])
        sheet.append(headers)
        for cell in sheet[3]:
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in rows:
            sheet.append(row)
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(headers)):
            for cell in row:
                cell.alignment = body_alignment
                cell.border = border
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
        for row_index in range(4, sheet.max_row + 1):
            sheet.row_dimensions[row_index].height = 72
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{sheet.max_row}"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_sheet(
        ws,
        f"专利写作反馈修复对照表（生成时间：{now}）",
        ["序号", "用户反馈 / 问题", "修复策略", "修复类型", "涉及文件", "关键实现", "效果 / 验证"],
        FIX_ROWS,
        [8, 34, 38, 20, 42, 52, 48],
    )
    write_sheet(
        summary,
        "文件变更摘要",
        ["文件", "作用", "本次改动摘要"],
        FILE_ROWS,
        [34, 32, 70],
    )
    wb.save(OUTPUT)
    load_workbook(OUTPUT, read_only=True).close()
    return True


CONTENT_TYPES = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>'''

RELS = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>'''

CORE_PROPS = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>Codex</dc:creator>
  <cp:lastModifiedBy>Codex</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">2026-06-23T00:00:00Z</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">2026-06-23T00:00:00Z</dcterms:modified>
</cp:coreProperties>'''

APP_PROPS = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Codex</Application>
  <DocSecurity>0</DocSecurity>
  <ScaleCrop>false</ScaleCrop>
  <HeadingPairs><vt:vector size="2" baseType="variant"><vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant><vt:variant><vt:i4>2</vt:i4></vt:variant></vt:vector></HeadingPairs>
  <TitlesOfParts><vt:vector size="2" baseType="lpstr"><vt:lpstr>修复对照</vt:lpstr><vt:lpstr>文件摘要</vt:lpstr></vt:vector></TitlesOfParts>
  <Company></Company>
  <LinksUpToDate>false</LinksUpToDate>
  <SharedDoc>false</SharedDoc>
  <HyperlinksChanged>false</HyperlinksChanged>
  <AppVersion>16.0000</AppVersion>
</Properties>'''

WORKBOOK = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="修复对照" sheetId="1" r:id="rId1"/>
    <sheet name="文件摘要" sheetId="2" r:id="rId2"/>
  </sheets>
</workbook>'''

WORKBOOK_RELS = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''

STYLES = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3">
    <font><sz val="11"/><name val="Arial"/></font>
    <font><b/><sz val="16"/><color rgb="FFFFFFFF"/><name val="Arial"/></font>
    <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Arial"/></font>
  </fonts>
  <fills count="4">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF145A4A"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF28715F"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border><left style="thin"><color rgb="FFD9E2DD"/></left><right style="thin"><color rgb="FFD9E2DD"/></right><top style="thin"><color rgb="FFD9E2DD"/></top><bottom style="thin"><color rgb="FFD9E2DD"/></bottom><diagonal/></border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment vertical="top" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''


if __name__ == "__main__":
    build_xlsx()
    print(OUTPUT)
