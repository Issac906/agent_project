"""Create a detailed Excel report for the patent-agent architecture optimization."""

from __future__ import annotations

from pathlib import Path
import sys


OPENPYXL_PATH = Path("/private/tmp/openpyxl_agent_report")
if OPENPYXL_PATH.exists():
    sys.path.insert(0, str(OPENPYXL_PATH))

from openpyxl import Workbook, load_workbook  # type: ignore[import-not-found]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-not-found]
from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]


OUTPUT = Path("outputs/patent_agent_skill_tool_optimization.xlsx")


SOLUTION_ROWS = [
    [
        1,
        "标题过长，仍生成“基于……的……”",
        "patent-writing 与 patent-quality-review 明确规定标题只保留发明对象，禁止“基于……的……”句式，算法、特征和模型名称必须移入发明内容。",
        "标题质量检查器识别“基于X的Y”和超过30字的标题；确定性修复函数直接把“基于X的Y”裁剪为“Y”。",
        "候选生成后先清洗标题；标题页、发明名称生成后再次检查；用户手动输入标题也走同一清洗工具。",
        "“基于VMD频带能量特征与时序网络的铝电解槽阳极效应早期预警方法”自动变为“铝电解槽阳极效应早期预警方法”。",
        "无论知识库是铝电解、油田或设备预测材料，标题均不再依赖模型自觉遵守，而由工具强制处理。",
    ],
    [
        2,
        "发明目的缺失",
        "patent-writing 把“发明目的”定义为发明内容的固定组成，位置位于关键创新点之后、拟解决技术问题之前。",
        "质量检查器检查发明内容是否包含“发明目的”；缺失时生成明确修复指令。",
        "章节生成 -> 检查缺项 -> Agent 只重写当前发明内容 -> 再检查，最多两轮。",
        "最终文档会出现独立的“发明目的”，说明提出该发明的业务目标和技术目标。",
        "不依赖知识库原文是否已经写出“目的”，Agent 会根据具体问题链提炼，但不编造事实。",
    ],
    [
        3,
        "保护范围缺失或只写一种方法",
        "patent-writing 和 patent-quality-review 要求保护范围覆盖方法、系统，以及适用时的装置/设备/存储介质，并围绕组合技术特征定义边界。",
        "质量检查器检查“方法”“系统”和“装置/设备/存储介质”是否覆盖；缺项即判不通过。",
        "生成保护范围 -> 检查保护形式 -> 自动补充缺失类别 -> 用户确认。",
        "最终文档新增独立“五、保护范围”，不会只有技术步骤而没有保护边界。",
        "面对不同知识库材料，工具按技术形态检查，而不是绑定某个行业或算法。",
    ],
    [
        4,
        "文档中重复出现“技术交底书”或重复总标题",
        "patent-writing 规定全文只保留一次文档类型表达；标题页负责文档类型，正文不重复创建总标题。",
        "最终文档检查器统计“技术交底书”出现次数，超过一次即报告重复标题。",
        "组装工具只拼接用户确认章节，不再额外插入候选标题；导出前执行全文复核。",
        "最终 Word/Markdown 不会在标题页之后再次出现第二个“技术交底书”。",
        "该检查针对最终成品结构，与知识库内容无关。",
    ],
    [
        5,
        "背景技术冗长，变成行业综述",
        "patent-writing 要求背景只保留2-3个本发明直接解决的行业问题；patent-quality-review 禁止泛泛行业历史和大段市场介绍。",
        "质量检查器检查背景长度、问题点数量和禁止表达；超过约1600字或不足2个问题点会失败。",
        "背景生成 -> 提取问题点 -> 检查长度和数量 -> 自动压缩/重写 -> 用户确认。",
        "背景部分呈现为清晰的问题列表，而不是长篇介绍行业发展。",
        "检查标准是“问题数量和聚焦程度”，可以适配任意行业素材。",
    ],
    [
        6,
        "把“填补技术空白”直接当作创新性",
        "patent-quality-review 禁止“技术空白一/二/三”“填补技术空白”“行业尚无”“业内尚无”等绝对化表述。",
        "质量检查器通过禁止词表检测这些表达，并要求改写为具体场景、数据、约束、评价或闭环机制上的不足。",
        "检测到绝对表述 -> 生成具体修复意见 -> Agent 仅重写当前背景章节 -> 再检查。",
        "最终背景会写“现有方案在某场景下缺少某能力”，不会直接宣称行业没有该技术。",
        "禁止词和修复方向固定在 Tool 中，不会因知识库措辞变化而失效。",
    ],
    [
        7,
        "背景问题与拟解决技术问题不对应",
        "patent-writing 定义问题链规则：背景提出a/b/c，发明内容必须按相同顺序解决a/b/c。",
        "质量检查器读取已经确认的背景章节，统计问题点，再检查“拟解决的技术问题”数量是否对应。",
        "先确认背景 -> 生成发明内容 -> 跨章节对照检查 -> 不对应时自动重写发明内容。",
        "最终文章形成“背景问题 -> 发明目的 -> 拟解决问题 -> 技术方案”的闭环逻辑。",
        "Tool 使用已确认前文章节作为上下文，适用于任何知识库主题。",
    ],
    [
        8,
        "关键创新点在后文再次单独重复",
        "patent-writing 把关键创新点固定在发明内容开头；patent-quality-review 禁止再生成“区别于现有技术的关键创新点”独立小节。",
        "质量检查器检查必要小节和重复小节名称；发现重复即要求合并。",
        "发明内容生成 -> 检查创新点位置 -> 删除后部重复内容 -> 再检查。",
        "创新点只出现一次，并在读者进入发明内容时优先展示。",
        "结构规则固定，不受素材章节组织方式影响。",
    ],
    [
        9,
        "有益效果出现没有来源的百分比、金额或精度",
        "patent-writing 和 patent-quality-review 规定：有证据才量化，无证据只写简洁定性效果并标记待验证。",
        "质量检查器识别“提升/降低/节省/达到 + 数值 + %、金额、倍数、时间”等表达，并与知识库和检索证据文本核对。",
        "检测量化结果 -> 查证据文本 -> 无相同依据则自动删除数值或改成定性说明 -> 再检查。",
        "最终有益效果不会出现模型凭空生成的“提升30%”“节省20万元”等数据。",
        "证据核验基于当前知识库材料，因此素材变化时仍按同一标准判断。",
    ],
    [
        10,
        "正文出现“目标行业”“目标对象”等泛化占位词",
        "patent-quality-review 要求正文使用当前知识库中的具体行业、设备、对象和应用场景；事实不足时写“待补充”。",
        "质量检查器检测“目标行业”“目标对象”“某行业”“某场景”等占位词。",
        "章节生成 -> 泛化词检测 -> Agent 根据知识库替换具体对象；知识库无依据则标记待补充。",
        "最终文档会体现当前知识库的真实技术对象，而不是通用模板文章。",
        "通过读取当前材料动态替换，避免模板绑定某个固定项目。",
    ],
    [
        11,
        "用户修改章节后可能绕过原有写作规则",
        "interactive-drafting 规定生成、重写、按意见修改和手动编辑都必须调用同一个质量审查流程。",
        "review_section() 是统一入口；手动编辑未通过时不会自动接受，而是展示问题。",
        "任何编辑动作 -> 质量检查 -> 通过才进入下一章；未通过则停留当前章节。",
        "用户修改后的内容仍符合标题、背景、问题对应、有益效果和保护范围要求。",
        "规则集中在一个 Tool 中，避免多个交互分支产生不同质量标准。",
    ],
    [
        12,
        "只靠提示词，模型可能忽略规则",
        "新增 patent-quality-review Skill，把写作标准定义为强制验收条件；agent-planning Skill 增加 review_patent_quality 工具步骤。",
        "新增 patent_quality_tool.py，以确定性规则给出分数、问题代码、具体原因和修复要求。",
        "生成 -> Tool 检查 -> Agent 按问题修复 -> Tool 复检 -> 用户确认 -> 全文复核 -> 导出。",
        "前端显示每章和最终文档的质量分、通过状态、问题及修复建议，改变能够直接在结果和界面中被看见。",
        "Skill 负责定义标准，Tool 负责强制执行，Agent 负责根据具体材料修复，三者职责分离。",
    ],
    [
        13,
        "多个Skill虽存在，但可能因上下文截断没有真正生效",
        "调整 Skill 加载顺序，把 patent-quality-review、patent-writing、interactive-drafting 放在前面；每个 Skill 设置独立预算。",
        "format_skills_for_prompt 总预算从12000提高到32000，并给每个Skill设置最低内容额度；Pi 调用显式传入质量Skill路径。",
        "启动 Agent -> 加载6个项目Skill -> 构建完整Skill上下文 -> 写作和修复均引用该上下文。",
        "质量、写作、交互、规划、素材评估、相似专利分析六个Skill都会实际进入Agent上下文。",
        "避免“文件存在但运行时没读到”的伪 Skill 架构。",
    ],
    [
        14,
        "最终文章如何证明应用确实完成了优化",
        "patent-quality-review 规定导出前必须执行全文质量门禁，并检查全部标准章节。",
        "review_document() 重新拆分最终文章，逐章复用相同检查规则，生成最终质量报告。",
        "全部章节确认 -> 组装全文 -> 全文复核 -> 记录质量结果 -> 导出Markdown/Word -> 写入历史记录。",
        "最终页面和历史记录保存质量分及未解决项；输出文章本身体现短标题、问题链、创新点前置、有证据效果和保护范围。",
        "不仅记录“代码改了”，而是对最终产物执行可验证验收。",
    ],
]


FLOW_ROWS = [
    [1, "加载 Skills", "加载 patent-quality-review、patent-writing、interactive-drafting 等6个项目Skill。", "完整规则进入Agent上下文。"],
    [2, "生成章节", "Agent根据当前知识库材料生成当前章节。", "文章内容适配当前行业和技术对象。"],
    [3, "确定性检查", "patent_quality_tool 按章节类型执行标题、背景、问题链、效果证据、保护范围检查。", "生成质量分和具体问题清单。"],
    [4, "自动修订", "未通过时，把问题清单和知识库证据交给Agent，只重写当前章节。", "解决具体问题，不重新生成无关章节。"],
    [5, "再次检查", "最多复检两轮，仍未通过则把剩余问题展示给用户。", "避免无限循环，同时保留人工判断。"],
    [6, "用户确认", "通过检查的章节展示给用户接受、重写、修改或手动编辑。", "Human-in-the-loop仍然保留。"],
    [7, "最终全文检查", "组装全文后再次逐章检查章节完整性和跨章节逻辑。", "防止手动编辑或组装阶段重新引入问题。"],
    [8, "导出与留痕", "导出Markdown/Word，并在前端和历史记录保存质量结果。", "导师和用户可看到优化是否真正落地。"],
]


def write_sheet(sheet, title: str, headers: list[str], rows: list[list[object]], widths: list[int], row_height: int) -> None:
    dark = "145A4A"
    green = "28715F"
    line = "D9E2DD"
    light = "F7FBF8"
    warning = "FFF8E8"

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = PatternFill("solid", fgColor=dark)
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.append([])
    sheet.append(headers)
    for cell in sheet[3]:
        cell.fill = PatternFill("solid", fgColor=green)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=dark))

    for row in rows:
        sheet.append(row)

    thin = Side(style="thin", color=line)
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(headers)),
        start=4,
    ):
        fill_color = light if row_index % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        row[0].alignment = Alignment(horizontal="center", vertical="top")

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(4, sheet.max_row + 1):
        sheet.row_dimensions[index].height = row_height

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{sheet.max_row}"


def build() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    solutions = workbook.active
    solutions.title = "问题与解决方案"
    flow = workbook.create_sheet("Skill-Tool执行闭环")

    write_sheet(
        solutions,
        "专利 Agent：Skill + Tool 架构级优化方案",
        ["序号", "要解决的问题", "Skill层具体改造", "Tool层具体改造", "实际执行机制", "最终文章中的体现", "跨素材稳定性"],
        SOLUTION_ROWS,
        [7, 28, 45, 48, 42, 43, 40],
        112,
    )
    write_sheet(
        flow,
        "Skill + Tool 质量闭环",
        ["步骤", "环节", "具体执行", "产生结果"],
        FLOW_ROWS,
        [8, 24, 78, 48],
        60,
    )

    workbook.save(OUTPUT)
    load_workbook(OUTPUT, read_only=True).close()
    print(OUTPUT)


if __name__ == "__main__":
    build()
